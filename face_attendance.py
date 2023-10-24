import cv2
import numpy as np
import face_recognition as fr
import os
from datetime import datetime
import openpyxl

# defining path of images folder
path = 'FACE-ATTENDANCE/persons-img'
images = []
className = []  # for name of images
myList = os.listdir(path)
# print(myList)

for cl in myList:
    imgCurr = cv2.imread(f'{path}/{cl}')
    images.append(imgCurr)
    className.append(os.path.splitext(cl)[0])

print(className)

# creating function for finding encodings in all img/faces


def find_encoding(images):
    encodingList = []
    for img in images:
        img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
        encode = fr.face_encodings(img)[0]
        encodingList.append(encode)

    return encodingList


encodeListKnown = find_encoding(images)
print("encoding complete")


# mark attendance function for marking the attendance in Excel file
def markAttendance(name):
    curr_time = datetime.now()
    date_string = curr_time.strftime('%H:%M:%S')

    try:
        # Try to open the Excel file, create it if it doesn't Exist
        wb = openpyxl.load_workbook('FACE-ATTENDANCE/attendance.xlsx')
    except FileNotFoundError:
        # If the Excel file doesnt exist, create a new workbook and a worksheet with headers
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['Name', 'Time'])
    else:
        # If the Excel file already exists, select the active worksheet
        ws = wb.active

    # Check if the name is already in the Excel file
    name_present = False
    for row in ws.iter_rows(values_only=True):
        if name in row:
            name_present = True
            break

    if name_present:
        print(f"{name} is already marked")
        return -1
    else:
        # Append a new row to the Excel file
        ws.append([name, date_string])
        wb.save('FACE-ATTENDANCE/attendance.xlsx')
        print(f"Attendance marked for {name} at {date_string}")


# starting webcam -> use '0' as an argument in cv2.VideoCapture() function, while using computer's default camera.
cap = cv2.VideoCapture(1)

while True:
    success, frame = cap.read()
    frame_rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)

    # finding faces in the current frame (ie. webcam)
    face_in_curr_frame = fr.face_locations(frame_rgb)
    encodes_curr_frame = fr.face_encodings(frame_rgb, face_in_curr_frame)

    for encode_face, face_loc in zip(encodes_curr_frame, face_in_curr_frame):
        matches = fr.compare_faces(encodeListKnown, encode_face)
        face_dist = fr.face_distance(encodeListKnown, encode_face)
        # print(face_dist)
        match_index = np.argmin(face_dist)

        if matches[match_index]:
            name = className[match_index].upper()
            print(name)

            # draww rectangle around the detected faces
            y1, x2, y2, x1 = face_loc
            cv2.rectangle(frame, (x1, y1), (x2, y2), (0, 255, 0), 2)
            cv2.rectangle(frame, (x1, y2 + 27), (x2, y2), (0, 255, 0), cv2.FILLED)
            cv2.putText(frame, name, (x1 + 7, y2 + 18), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (255, 255, 255), 2)

            # now mark attendance of recognized-faces
            result = markAttendance(name)
            if (result == -1):
                cv2.putText(frame, "Attendance Marked!", (140, 30), cv2.FONT_HERSHEY_SIMPLEX, 1, (255, 0, 255), 2)

        # if face doesn't match, write UNKNOWN
        else:
            y1, x2, y2, x1 = face_loc
            cv2.rectangle(frame, (x1, y1), (x2, y2), (0, 0, 255), 2)
            cv2.rectangle(frame, (x1, y2 + 27), (x2, y2), (0, 0, 255), cv2.FILLED)
            cv2.putText(frame, "UNKNOWN", (x1 + 7, y2 + 18), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (255, 255, 255), 2)
            print("Unknown face detected")

    # showing the output window, and stop when 'esc' is pressed
    cv2.imshow("output", frame)
    if cv2.waitKey(1) == 27:
        break
